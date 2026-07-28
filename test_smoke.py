# -*- coding: utf-8 -*-
"""Ağ gerektirmeyen duman testi."""
import io, os, tempfile
import openpyxl
import backends as B
from cache import CountCache

# 1) sorgu kurulumu
syn = {"Citrate": ["citrate", "citric acid"]}
pm = B.PubMedBackend()
ep = B.EuropePMCBackend()
print("PubMed :", pm.build_query("Citrate", "urine", syn, 2010, 2026))
print("EPMC   :", ep.build_query("Citrate", "urine", syn, 2010, 2026))
print("EPMC bilinmeyen metabolit:", ep.build_query("Taurine", "csf", syn))
print("gecikme (anahtarsız/anahtarlı):", pm.limiter.min_interval,
      B.PubMedBackend(api_key="x").limiter.min_interval)
print("matris tahmini:", [(s, B.guess_matrix_key(s)) for s in
      ["1. Serum-Plazma", "2. Idrar-Urine", "5. BOS-CSF", "4. Feces", "Sheet9"]])

# 2) önbellek
path = os.path.join(tempfile.mkdtemp(), "c.sqlite")
c = CountCache(path)
c.put("PubMed", "q1", 42)
assert c.get("PubMed", "q1")[0] == 42
assert c.get("PubMed", "yok") is None
assert c.get("PubMed", "q1", max_age_days=0)[0] == 42        # bugün çekildi
c.put("PubMed", "q2", 7, fetched_at="2020-01-01T00:00:00")
assert c.get("PubMed", "q2", max_age_days=30) is None        # bayat
assert c.get("PubMed", "q2")[0] == 7                         # TTL yoksa geçerli
blob = c.export_csv()
c2 = CountCache(os.path.join(tempfile.mkdtemp(), "d.sqlite"))
assert c2.import_csv(blob) == 2 and c2.total() == 2
print("önbellek: OK", c.stats())

# 3) çalışma kitabı yazımı (sahte arka uçlarla, ağ yok)
wb = openpyxl.Workbook()
ws = wb.active; ws.title = "2. Urine"
ws.cell(4, 3, "Metabolite")
for i, m in enumerate(["Citrate", "Acetate", "Formate", "Taurine"], start=5):
    ws.cell(i, 1, i - 4); ws.cell(i, 3, m)
tmp = os.path.join(tempfile.mkdtemp(), "wb.xlsx"); wb.save(tmp)

class Stub:
    def __init__(self, name, table):
        self.name = name; self.table = table
        self.limiter = type("L", (), {"min_interval": 0.0})()
    def build_query(self, met, mkey, syn, y0=None, y1=None):
        return f"{self.name}:{met}:{mkey}"
    def count(self, q):
        return self.table[q.split(":")[1]], None
    def describe(self): return self.name

a = Stub("PubMed",     {"Citrate": 500, "Acetate": 300, "Formate": 100, "Taurine": 50})
b = Stub("Europe PMC", {"Citrate": 900, "Acetate": 250, "Formate": 400, "Taurine": 40})

import app  # bare mode
wb2 = openpyxl.load_workbook(tmp)
res = app.run_counts(wb2, {"2. Urine": "urine"}, syn, [a, b], a,
                     {"from_year": None, "to_year": None, "sort": True,
                      "pct_value": True, "ttl": None}, None)
df = res["previews"]["2. Urine"]
print(df.to_string(index=False))
ws2 = wb2["2. Urine"]
print("E sütunu (birincil sayım):", [ws2.cell(r, 5).value for r in range(5, 9)])
print("F sütunu (pay)          :", [round(ws2.cell(r, 6).value, 4) for r in range(5, 9)])
print("I..L başlıkları         :", [ws2.cell(4, c).value for c in range(9, 13)])
print("I..L ilk satır          :", [ws2.cell(5, c).value for c in range(9, 13)])
print("Spearman ρ:", app.spearman(df["PubMed sayı"], df["Europe PMC sayı"]))
print("ilk 2 örtüşme:", app.topn_overlap(df, "PubMed sayı", "Europe PMC sayı", 2))
assert abs(sum(ws2.cell(r, 6).value for r in range(5, 9)) - 1.0) < 1e-9
print("\nTÜM TESTLER GEÇTİ")
