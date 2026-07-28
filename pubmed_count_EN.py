#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
app.py — PubMed / Europe PMC yayın sayımı (metabolit × biyoakışkan)

Arayüz dört dilde çalışır (tr, en, de, fr); metinler locales/ altındadır.

Çalıştırma:
    pip install -r requirements.txt
    streamlit run app.py
"""

import csv
import datetime as _dt
import io
import json
import os
import re

import openpyxl
import pandas as pd
import streamlit as st
from openpyxl.styles import Alignment, Font

import i18n
from backends import MATRIX_KEYS, EuropePMCBackend, PubMedBackend, guess_matrix_key
from branding import find_logo, header, page_icon
from cache import CountCache
from i18n import t

# ----------------------------------------------------------------- sayfa düzeni
SHEET_COLS = {"metabolite_en": 3, "count": 5, "query": 8}
HEADER_ROW = 4  # veri satırları 5'ten başlar
EXTRA_COL = 9   # karşılaştırma sütunları buradan itibaren yazılır

MATRIX_SHEETS = [
    "1. Serum-Plazma", "2. Idrar", "3. Salya", "4. Gayta", "5. BOS",
    "1. Serum-Plasma", "2. Urine", "3. Saliva", "4. Feces", "5. CSF",
    "2. Idrar-Urine", "3. Salya-Saliva", "4. Gayta-Feces", "5. BOS-CSF",
]

SYNONYMS = {
    "Lactic acid":        ["lactic acid", "lactate"],
    "Glutamic acid":      ["glutamic acid", "glutamate"],
    "Aspartic acid":      ["aspartic acid", "aspartate"],
    "Uric acid":          ["uric acid", "urate"],
    "Citrate":            ["citrate", "citric acid"],
    "Succinate":          ["succinate", "succinic acid"],
    "Formate":            ["formate", "formic acid"],
    "Acetate":            ["acetate", "acetic acid"],
    "Pyruvate":           ["pyruvate", "pyruvic acid"],
    "Butyrate":           ["butyrate", "butyric acid"],
    "Propionate":         ["propionate", "propionic acid"],
    "Isovalerate":        ["isovalerate", "isovaleric acid"],
    "Isobutyrate":        ["isobutyrate", "isobutyric acid"],
    "Valerate":           ["valerate", "valeric acid"],
    "Oxalate":            ["oxalate", "oxalic acid"],
    "Hippurate":          ["hippurate", "hippuric acid"],
    "2-Oxoglutarate":     ["2-oxoglutarate", "alpha-ketoglutarate", "2-ketoglutarate"],
    "3-Hydroxybutyrate":  ["3-hydroxybutyrate", "beta-hydroxybutyrate", "3-hydroxybutyric acid"],
    "Methylmalonic acid": ["methylmalonic acid", "methylmalonate"],
    "Orotic acid":        ["orotic acid", "orotate"],
    "Pyroglutamate":      ["pyroglutamate", "pyroglutamic acid", "5-oxoproline"],
    "Trimethylamine N-oxide (TMAO)": ["trimethylamine N-oxide", "TMAO"],
    "GABA (4-aminobutyrate)": ["GABA", "gamma-aminobutyric acid", "4-aminobutyrate"],
    "Lysophosphatidylcholines (LPC)": ["lysophosphatidylcholine", "LPC"],
    "N-Acetylaspartate (NAA)": ["N-acetylaspartate", "NAA"],
    "5-Hydroxyindoleacetic acid": ["5-hydroxyindoleacetic acid", "5-HIAA"],
    "Homovanillic acid (HVA)": ["homovanillic acid", "HVA"],
    "Hexadecanoic acid (palmitic)": ["palmitic acid", "hexadecanoic acid"],
    "Carnitine":          ["carnitine", "L-carnitine"],
    "Acetylcarnitine":    ["acetylcarnitine", "acetyl-L-carnitine"],
}


# --------------------------------------------------------------------- yardımcı
def secret(key, default=""):
    try:
        return st.secrets.get(key, default)
    except Exception:
        return default


@st.cache_resource(show_spinner=False)
def get_cache(path: str) -> CountCache:
    return CountCache(path)


def sheet_row_count(ws) -> int:
    n, r = 0, HEADER_ROW + 1
    while ws.cell(r, SHEET_COLS["metabolite_en"]).value:
        n += 1
        r += 1
    return n


def spearman(a: pd.Series, b: pd.Series):
    """Sıra korelasyonu — sıralara uygulanan Pearson. scipy gerektirmez."""
    ok = a.notna() & b.notna()
    if ok.sum() < 3:
        return None
    return float(a[ok].rank().corr(b[ok].rank()))


def topn_overlap(df, col_a, col_b, n):
    top_a = set(df.nlargest(n, col_a)["metabolite"])
    top_b = set(df.nlargest(n, col_b)["metabolite"])
    return len(top_a & top_b), sorted(top_a ^ top_b)


def backend_label(backend) -> str:
    """Arka ucun durumunu seçili dilde anlatır."""
    info = backend.info()
    if info["name"] == "PubMed":
        return t("backend.pubmed_keyed") if info.get("keyed") else t("backend.pubmed_nokey")
    return t(
        "backend.epmc",
        state=t("state.on") if info.get("synonym") else t("state.off"),
        medline=t("backend.medline_suffix") if info.get("medline") else "",
    )


# Önizleme tablosunun sütun adları dilden bağımsız tutulur; çeviri yalnızca
# gösterim anında uygulanır. Aksi hâlde koşudan sonra dil değiştirilince
# session_state'teki tablo ile arama anahtarları uyuşmaz.
def col_count(name):  return f"{name}|count"
def col_rank(name):   return f"{name}|rank"


def display_df(df: pd.DataFrame, names) -> pd.DataFrame:
    mapping = {"metabolite": t("col.metabolite"), "delta_rank": t("col.delta")}
    for name in names:
        mapping[col_count(name)] = f"{name} {t('col.count')}"
        mapping[col_rank(name)] = f"{name} {t('col.rank')}"
    return df.rename(columns=mapping)


# ------------------------------------------------------------------ ana işleyiş
def run_counts(wb, sheet_map, synonyms, backends, primary, opts, cache):
    """Seçili sayfaları işler, çalışma kitabını yerinde günceller."""
    stamp = _dt.datetime.now().isoformat(timespec="seconds")
    log_rows, previews, errors = [], {}, []
    hits = misses = 0

    total = sum(sheet_row_count(wb[s]) for s in sheet_map) * len(backends)
    done = 0
    bar = st.progress(0.0)
    status = st.empty()

    for sheet, mkey in sheet_map.items():
        ws = wb[sheet]
        records = []

        r = HEADER_ROW + 1
        while ws.cell(r, SHEET_COLS["metabolite_en"]).value:
            met = str(ws.cell(r, SHEET_COLS["metabolite_en"]).value).strip()
            ncols = max(ws.max_column, SHEET_COLS["query"])
            rec = {
                "met": met,
                "vals": [ws.cell(r, c).value for c in range(1, ncols + 1)],
                "counts": {}, "queries": {},
            }

            for b in backends:
                q = b.build_query(met, mkey, synonyms,
                                  opts["from_year"], opts["to_year"])
                cached = cache.get(b.name, q, opts["ttl"]) if cache else None
                if cached is not None:
                    cnt, fetched = cached
                    hits += 1
                    from_cache = True
                else:
                    cnt, err = b.count(q)
                    from_cache = False
                    misses += 1
                    fetched = stamp
                    if err:
                        errors.append(f"**{sheet} / {met}** — {b.name}: {err}")
                    elif cache:
                        fetched = cache.put(b.name, q, cnt)

                rec["counts"][b.name] = cnt
                rec["queries"][b.name] = q
                log_rows.append([stamp, sheet, met, b.name, cnt,
                                 "yes" if from_cache else "no", fetched, q])

                done += 1
                bar.progress(min(done / total, 1.0) if total else 1.0)
                status.write(t(
                    "progress.line", sheet=sheet, metabolite=met, backend=b.name,
                    count=cnt if cnt is not None else t("progress.error"),
                    cached=t("progress.cached") if from_cache else "",
                    done=done, total=total,
                ))

            records.append(rec)
            r += 1

        previews[sheet] = _finalise_sheet(ws, records, backends, primary, opts)

    bar.empty()
    status.empty()
    return {"log": log_rows, "previews": previews, "errors": errors,
            "stamp": stamp, "hits": hits, "misses": misses}


def _finalise_sheet(ws, records, backends, primary, opts):
    """Satırları sıralar, çalışma sayfasına yazar, önizleme tablosu döndürür."""
    if not records:
        return pd.DataFrame()

    if opts["sort"]:
        records.sort(key=lambda rec: (-1 if rec["counts"].get(primary.name) is None
                                      else rec["counts"][primary.name]), reverse=True)

    df = pd.DataFrame({
        "metabolite": [rec["met"] for rec in records],
        **{col_count(b.name): [rec["counts"].get(b.name) for rec in records]
           for b in backends},
    })
    for b in backends:
        df[col_rank(b.name)] = df[col_count(b.name)].rank(
            ascending=False, method="min").astype("Int64")
    if len(backends) == 2:
        first, second = backends[0].name, backends[1].name
        df["delta_rank"] = (df[col_rank(first)] - df[col_rank(second)]).abs()

    total_cnt = sum(v for v in df[col_count(primary.name)] if pd.notna(v))
    last = HEADER_ROW + len(records)
    secondary = [b for b in backends if b.name != primary.name]

    if secondary:
        sec = secondary[0].name
        labels = [f"{sec} {t('col.count')}", f"{sec} {t('col.rank')}",
                  f"{primary.name} {t('col.rank')}", t("col.delta")]
        for offset, label in enumerate(labels):
            cell = ws.cell(HEADER_ROW, EXTRA_COL + offset, value=label)
            cell.font = Font(name="Arial", size=9, bold=True)
            cell.alignment = Alignment(vertical="center", horizontal="center",
                                       wrap_text=True)

    for i, rec in enumerate(records, start=1):
        rr = HEADER_ROW + i
        vals = rec["vals"]
        vals[0] = i
        vals[SHEET_COLS["count"] - 1] = rec["counts"].get(primary.name)
        vals[SHEET_COLS["query"] - 1] = rec["queries"].get(primary.name)

        for c, v in enumerate(vals, start=1):
            cell = ws.cell(rr, c, value=v)
            cell.font = Font(name="Arial", size=9)
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=(c in (2, 3, 4, 7, 8)),
                horizontal="center" if c in (1, 5, 6) else None,
            )

        cnt = rec["counts"].get(primary.name)
        share = (cnt / total_cnt) if (cnt is not None and total_cnt) else None
        if opts["pct_value"]:
            ws.cell(rr, 6, value=share)
        else:
            ws.cell(rr, 6,
                    value=f'=IFERROR(E{rr}/SUM($E${HEADER_ROW + 1}:$E${last}),"")')
        ws.cell(rr, 6).number_format = "0.0%"

        if secondary:
            sec = secondary[0].name
            row = df.iloc[i - 1]
            for offset, v in enumerate([
                rec["counts"].get(sec),
                None if pd.isna(row[col_rank(sec)]) else int(row[col_rank(sec)]),
                None if pd.isna(row[col_rank(primary.name)])
                else int(row[col_rank(primary.name)]),
                None if pd.isna(row.get("delta_rank")) else int(row["delta_rank"]),
            ]):
                cell = ws.cell(rr, EXTRA_COL + offset, value=v)
                cell.font = Font(name="Arial", size=9)
                cell.alignment = Alignment(vertical="top", horizontal="center")

    df.insert(0, "#", range(1, len(df) + 1))
    return df


# ------------------------------------------------------------------------ arayüz
i18n.current()  # dili set_page_config'ten önce çöz
_logo = find_logo()
st.set_page_config(
    page_title=t("app.title"),
    page_icon=page_icon(_logo),
    layout="wide",
)
header(t("app.title"), t("app.caption"), logo=_logo)

ss = st.session_state
for k, v in {"out_xlsx": None, "out_log": None, "res": None,
             "out_name": "counted.xlsx", "backend_names": []}.items():
    ss.setdefault(k, v)

with st.sidebar:
    i18n.selector()
    st.divider()

    st.header(t("source.header"))
    modes = ["PubMed", "Europe PMC", t("source.mode.both")]
    mode = st.radio(t("source.mode_label"), modes, index=0,
                    label_visibility="collapsed", key="mode")
    both = mode == modes[2]
    use_pm = mode != "Europe PMC"
    use_ep = mode != "PubMed"

    primary_name = "PubMed"
    if both:
        primary_name = st.selectbox(t("source.primary"), ["PubMed", "Europe PMC"],
                                    help=t("source.primary_help"), key="primary")

    if use_pm:
        st.subheader("PubMed")
        api_key = st.text_input(t("pubmed.api_key"), type="password",
                                value=secret("NCBI_API_KEY"), key="pm_key")
        email = st.text_input(t("pubmed.email"), key="pm_mail")
        tool = st.text_input(t("pubmed.tool"), value="met4metab-count", key="pm_tool")
        pm_delay = st.number_input(t("pubmed.delay"), 0.05, 5.0,
                                   0.11 if api_key else 0.40, step=0.01,
                                   format="%.2f", help=t("pubmed.delay_help"),
                                   key="pm_delay")
        if not api_key:
            st.caption(t("pubmed.nokey"))
    else:
        api_key = email = tool = ""
        pm_delay = 0.40

    if use_ep:
        st.subheader("Europe PMC")
        st.caption(t("epmc.note"))
        epmc_synonym = st.checkbox(t("epmc.synonym"), value=False,
                                   help=t("epmc.synonym_help"), key="ep_syn")
        epmc_medline = st.checkbox(t("epmc.medline"), value=True,
                                   help=t("epmc.medline_help"), key="ep_med")
        human_filter = st.text_input(t("epmc.human_filter"), value='MESH_TERMS:"Humans"',
                                     help=t("epmc.human_filter_help"), key="ep_hf")
        epmc_delay = st.number_input(t("epmc.delay"), 0.05, 5.0, 0.20, step=0.01,
                                     format="%.2f", key="ep_delay")
    else:
        epmc_synonym, epmc_medline, epmc_delay = False, True, 0.20
        human_filter = 'MESH_TERMS:"Humans"'

    st.divider()
    st.header(t("query.header"))
    use_years = st.checkbox(t("query.use_years"), key="use_years")
    from_year = to_year = None
    if use_years:
        c1, c2 = st.columns(2)
        from_year = c1.number_input(t("query.from"), 1800, 2100, 2010, step=1, key="y0")
        to_year = c2.number_input(t("query.to"), 1800, 2100,
                                  _dt.date.today().year, step=1, key="y1")

    st.divider()
    st.header(t("cache.header"))
    use_cache = st.checkbox(t("cache.use"), value=True, key="use_cache")
    cache_path = st.text_input(t("cache.file"),
                               value=os.environ.get("COUNT_CACHE", "pubmed_cache.sqlite"),
                               key="cache_path")
    ttl = st.number_input(t("cache.ttl"), 0, 3650, 30, step=1,
                          help=t("cache.ttl_help"), key="ttl")
    cache_obj = get_cache(cache_path) if use_cache else None
    if cache_obj:
        st.caption(t("cache.records", n=cache_obj.total()))
        if cache_obj.total():
            stats = pd.DataFrame(cache_obj.stats()).rename(columns={
                "backend": t("cache.col.backend"), "records": t("cache.col.records"),
                "oldest": t("cache.col.oldest"), "newest": t("cache.col.newest"),
            })
            st.dataframe(stats, hide_index=True)
        cc1, cc2 = st.columns(2)
        if cc1.button(t("cache.clear"), use_container_width=True, key="cache_clear"):
            st.toast(t("cache.cleared", n=cache_obj.clear()))
            st.rerun()
        cc2.download_button(t("cache.export"), cache_obj.export_csv(),
                            file_name="count_cache.csv", mime="text/csv",
                            use_container_width=True, key="cache_export")
        imp = st.file_uploader(t("cache.import_label"), type=["csv"], key="cache_import")
        if imp is not None and st.button(t("cache.import"), use_container_width=True,
                                         key="cache_import_btn"):
            st.toast(t("cache.imported", n=cache_obj.import_csv(imp.getvalue())))
            st.rerun()
        st.caption(t("cache.cloud_note"))

    st.divider()
    st.header(t("output.header"))
    do_sort = st.checkbox(t("output.sort"), value=True, key="sort")
    pct_value = st.checkbox(t("output.pct_value"), value=True,
                            help=t("output.pct_help"), key="pct")
    topn = st.number_input(t("output.topn"), 5, 200, 20, step=5, key="topn")

# arka uçları kur
backends = []
if use_pm:
    backends.append(PubMedBackend(api_key or None, email or None,
                                  tool or "met4metab-count", pm_delay))
if use_ep:
    backends.append(EuropePMCBackend(epmc_delay, epmc_synonym, epmc_medline,
                                     human_filter, email or None))
primary = next((b for b in backends if b.name == primary_name), backends[0])

st.info(" · ".join(backend_label(b) for b in backends) +
        (t("info.primary", name=primary.name) if both else ""))

with st.expander(t("syn.expander")):
    syn_text = st.text_area("SYNONYMS", json.dumps(SYNONYMS, ensure_ascii=False, indent=2),
                            height=240, label_visibility="collapsed", key="syn")
    try:
        synonyms = json.loads(syn_text)
        st.caption(t("syn.count", n=len(synonyms)))
    except json.JSONDecodeError as exc:
        synonyms = SYNONYMS
        st.error(t("syn.error", err=exc))

uploaded = st.file_uploader(t("upload.label"), type=["xlsx"], key="workbook")

if uploaded is None:
    st.caption(t("upload.hint"))
else:
    raw = uploaded.getvalue()
    probe = openpyxl.load_workbook(io.BytesIO(raw))
    auto = [s for s in dict.fromkeys(MATRIX_SHEETS) if s in probe.sheetnames]
    chosen = st.multiselect(t("sheets.label"), probe.sheetnames,
                            default=auto or probe.sheetnames, key="sheets")

    sheet_map = {}
    if chosen:
        with st.expander(t("matrix.expander"), expanded=not auto):
            for s in chosen:
                pick = st.selectbox(s, MATRIX_KEYS,
                                    index=MATRIX_KEYS.index(guess_matrix_key(s)),
                                    format_func=lambda k: t(f"matrix.{k}"),
                                    key=f"mx_{s}")
                sheet_map[s] = pick

        rows = sum(sheet_row_count(probe[s]) for s in chosen)
        eta = rows * sum(b.limiter.min_interval for b in backends)
        st.caption(t("eta", sheets=len(chosen), queries=rows * len(backends),
                     minutes=f"{eta / 60:.1f}"))

    if st.button(t("run.button"), type="primary", disabled=not chosen, key="run"):
        wb = openpyxl.load_workbook(io.BytesIO(raw))
        res = run_counts(
            wb, sheet_map, synonyms, backends, primary,
            {"from_year": from_year, "to_year": to_year, "sort": do_sort,
             "pct_value": pct_value, "ttl": (None if ttl == 0 else int(ttl))},
            cache_obj,
        )

        buf = io.BytesIO()
        wb.save(buf)
        ss.out_xlsx = buf.getvalue()
        ss.out_name = re.sub(r"\.xlsx$", "_counted.xlsx", uploaded.name)

        sio = io.StringIO()
        w = csv.writer(sio)
        # log başlıkları bilinçli olarak İngilizce: makine tarafından okunan,
        # dilden bağımsız bir yeniden üretilebilirlik kaydı
        w.writerow(["run_timestamp", "sheet", "metabolite", "backend",
                    "count", "from_cache", "fetched_at", "query"])
        w.writerows(res["log"])
        ss.out_log = sio.getvalue().encode("utf-8-sig")
        ss.res = res
        ss.backend_names = [b.name for b in backends]

# ------------------------------------------------------------------- sonuçlar
res = ss.res
if res and ss.out_xlsx:
    st.success(t("result.success", stamp=res["stamp"],
                 requests=res["misses"], cached=res["hits"]))
    if res["errors"]:
        with st.expander(t("result.errors", n=len(res["errors"]))):
            for e in res["errors"]:
                st.write("- " + e)

    c1, c2 = st.columns(2)
    c1.download_button(t("dl.workbook"), ss.out_xlsx, file_name=ss.out_name,
                       mime="application/vnd.openxmlformats-officedocument."
                            "spreadsheetml.sheet", use_container_width=True,
                       key="dl_wb")
    c2.download_button(t("dl.log"), ss.out_log, file_name="count_log.csv",
                       mime="text/csv", use_container_width=True, key="dl_log")

    names = ss.backend_names
    for tab, (sheet, df) in zip(st.tabs(list(res["previews"])),
                                res["previews"].items()):
        with tab:
            if df.empty:
                st.warning(t("sheet.empty"))
                continue
            shown = display_df(df, names)
            st.dataframe(shown, use_container_width=True, hide_index=True)

            if len(names) == 2:
                a, b = names
                rho = spearman(df[col_count(a)], df[col_count(b)])
                n = min(int(topn), len(df))
                shared, diff = topn_overlap(df, col_count(a), col_count(b), n)

                st.subheader(t("cmp.header"))
                m1, m2, m3 = st.columns(3)
                m1.metric(t("cmp.spearman"), f"{rho:.3f}" if rho is not None else "—")
                m2.metric(t("cmp.overlap", n=n), f"{shared}/{n}")
                m3.metric(t("cmp.maxdelta"),
                          int(df["delta_rank"].max())
                          if df["delta_rank"].notna().any() else "—")

                st.scatter_chart(shown, x=f"{a} {t('col.rank')}",
                                 y=f"{b} {t('col.rank')}")
                if diff:
                    st.caption(t("cmp.only_one", n=n, items=", ".join(diff)))
                st.caption(t("cmp.note"))
            else:
                col = col_count(names[0])
                if df[col].notna().any():
                    st.bar_chart(df.head(20).set_index("metabolite")[col])

# --------------------------------------------------------------- tek sorgu testi
with st.expander(t("test.expander")):
    tb_name = st.selectbox(t("test.backend"), [b.name for b in backends],
                           key="test_backend")
    tb = next(b for b in backends if b.name == tb_name)
    default_q = tb.build_query("Citrate", "urine", SYNONYMS, from_year, to_year)
    tq = st.text_area(t("test.query"), value=default_q, height=110, key="test_query")
    if st.button(t("test.run"), key="test_run"):
        cnt, err = tb.count(tq)
        if err:
            st.error(err)
        else:
            st.metric(t("test.result"), f"{cnt:,}".replace(",", " "))

st.caption(t("footer"))
